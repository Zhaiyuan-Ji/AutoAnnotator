import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os


sentiment_path = r'C:\Users\21888\Desktop\test\finetune_toxic.xlsx'
intermediate_path = r'C:\Users\21888\Desktop\test\66.xlsx'
final_result_path = r'C:\Users\21888\Desktop\test\final_result.xlsx'

try:

    df_sentiment = pd.read_excel(sentiment_path)


    if os.path.exists(intermediate_path):
        df_intermediate = pd.read_excel(intermediate_path)
    else:
        df_intermediate = pd.DataFrame(columns=['text', 'label'])


    required_columns = ['text', 'label']
    for df in [df_sentiment, df_intermediate]:
        for col in required_columns:
            if col not in df.columns:
                raise ValueError(f"need {col}")

    df_sentiment = df_sentiment.drop_duplicates(subset='text', keep='last')


    text_to_label = df_sentiment.set_index('text')['label'].to_dict()


    empty_label_mask = df_intermediate['label'].isna() | (df_intermediate['label'] == '')


    df_intermediate.loc[empty_label_mask, 'label'] = df_intermediate.loc[empty_label_mask, 'text'].map(text_to_label)

    df_intermediate.to_excel(intermediate_path, index=False)

    if os.path.exists(final_result_path):
        book = load_workbook(final_result_path)
        sheet = book.active
        for r_idx, row in enumerate(dataframe_to_rows(df_intermediate, index=False, header=False), start=sheet.max_row + 1):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
        book.save(final_result_path)
    else:
        df_intermediate.to_excel(final_result_path, index=False)

except FileNotFoundError:
    print("no file")
except ValueError as ve:
    print(ve)
except Exception as e:
    print(f"error: {e}")